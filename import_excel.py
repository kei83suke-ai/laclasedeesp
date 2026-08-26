"""Build the offline Spanish learning dataset from the source workbook.

The workbook is treated as source material, not as the runtime database.  This
builder normalizes display text, assigns deterministic IDs, separates simple
verbs from verb phrases, and joins all 438 verb records to conjugation data.
External sources are used only during an audit/build step; the generated app
contains no network dependency.
"""

from __future__ import annotations

import hashlib
import json
import re
import unicodedata
from pathlib import Path

import openpyxl


PROJECT_DIR = Path(__file__).resolve().parent
ROOT_DIR = PROJECT_DIR.parent.parent
EXCEL_DIR = ROOT_DIR / "01単語帳"
EXCEL_CANDIDATES = sorted(EXCEL_DIR.glob("*完成版.xlsx"))
EXCEL_PATH = EXCEL_CANDIDATES[0] if EXCEL_CANDIDATES else EXCEL_DIR / "スペイン語単語　完成版.xlsx"
JS_PATH = PROJECT_DIR / "js" / "data.js"
DATA_DIR = PROJECT_DIR / "data"
CEFR_PATH = DATA_DIR / "cefr_levels.json"
AUDIT_MANIFEST_PATH = DATA_DIR / "audit_manifest.json"

WORD_SHEETS = {
    "名詞": ("nouns", "noun"),
    "動詞": ("verbs", "verb"),
    "形容詞": ("adjectives", "adjective"),
    "副詞": ("adverbs", "adverb"),
    "その他": ("prepositions", "other"),
    "フレーズ": ("phrases", "phrase"),
    "挨拶": ("phrases", "phrase"),
}

CONJUGATION_SHEETS = {
    "現在・過去分詞": "現在・過去分詞",
    "現在形": "現在形 (Presente)",
    "点過去": "点過去 (Indefinido)",
    "線過去": "線過去 (Imperfecto)",
    "未来形": "未来形 (Futuro)",
    "接続法現在": "接続法現在 (Subjuntivo)",
    "命令形": "命令形 (Imperativo)",
}

PERSONS = ("yo", "tu", "el/ella", "nosotros", "ellos")
REFLEXIVE_PRONOUNS = {
    "yo": "me",
    "tu": "te",
    "el/ella": "se",
    "nosotros": "nos",
    "ellos": "se",
}


def normalize_text(value: object) -> str:
    return re.sub(r"\s+", " ", unicodedata.normalize("NFKC", str(value or "")).strip())


def normalize_lexeme(value: object) -> str:
    text = normalize_text(value)
    return re.sub(r"\s*/\s*", "/", text)


def search_key(value: object) -> str:
    text = normalize_lexeme(value).casefold()
    text = unicodedata.normalize("NFD", text)
    return "".join(char for char in text if not unicodedata.combining(char))


def dedup_key(value: object) -> str:
    """Accent-preserving key: cuándo and cuando are different lemmas."""
    return normalize_lexeme(value).casefold()


def parse_region(text: str) -> tuple[str, str]:
    match = re.search(r"\((LA|Am\.?|España|Esp\.?)\)", text, flags=re.IGNORECASE)
    if not match:
        return text, ""
    region = match.group(1).upper().replace(".", "")
    canonical = normalize_text(text[: match.start()] + text[match.end() :])
    return canonical, region


def slugify(value: str) -> str:
    text = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-zA-Z0-9]+", "-", text.lower()).strip("-")
    return text or "item"


def stable_id(prefix: str, canonical_es: str, region: str, en: str, ja: str, used: set[str]) -> str:
    base = f"{prefix}:{slugify(canonical_es)}"
    if region:
        base += f":{slugify(region)}"
    if base not in used:
        used.add(base)
        return base
    digest = hashlib.sha1(f"{prefix}|{canonical_es}|{region}|{en}|{ja}".encode("utf-8")).hexdigest()[:8]
    candidate = f"{base}:{digest}"
    used.add(candidate)
    return candidate


def is_header(es: str, ja: str) -> bool:
    # 「el japonés / 日本語」のような正規の語彙をヘッダーと誤認しない。
    # 各語彙シートは2行目から読み込むため、ここでは明らかな見出し語だけを除外する。
    return es in {"スペイン語", "動詞", "名詞", "形容詞", "副詞"}


def word_key_for_sheet(sheet_name: str) -> tuple[str | None, str | None]:
    clean_name = sheet_name.strip()
    if clean_name in CONJUGATION_SHEETS:
        return None, None
    if "動詞" in clean_name and any(token in clean_name for token in ("分詞", "Presente", "Indefinido", "Imperfecto", "Futuro", "Subjuntivo", "現在", "過去", "未来", "接続", "命令")):
        return None, None
    for marker, value in WORD_SHEETS.items():
        if marker in clean_name:
            return value
    return None, None


def load_cefr_levels() -> dict[str, dict]:
    if not CEFR_PATH.exists():
        return {}
    try:
        value = json.loads(CEFR_PATH.read_text(encoding="utf-8"))
        return value if isinstance(value, dict) else {}
    except (OSError, json.JSONDecodeError):
        return {}


def audit_fields(entry: dict, cefr_levels: dict[str, dict]) -> dict:
    key = f"{entry['partOfSpeech']}|{search_key(entry['canonicalEs'])}|{search_key(entry.get('region', ''))}"
    cefr = cefr_levels.get(key, {})
    level = cefr.get("level")
    return {
        "cefrLevel": level,
        "cefrStatus": "pcic_matched" if level else "unmatched",
        "cefrSource": cefr.get("source"),
        "translationStatus": "workbook_retained",
        "translationSource": "source_workbook",
        "lexemeStatus": "normalized",
    }


def deduplicate_entries(entries: list[dict]) -> tuple[list[dict], list[dict]]:
    """Merge identical lemmas while preserving every source-row alias."""
    merged: dict[str, dict] = {}
    duplicate_groups = []
    for entry in entries:
        key = "|".join([
            entry.get("partOfSpeech", ""),
            dedup_key(entry.get("canonicalEs", "")),
            dedup_key(entry.get("region", "") or ""),
        ])
        existing = merged.get(key)
        if existing is None:
            existing = dict(entry)
            existing["legacyIds"] = list(entry.get("legacyIds", []))
            existing["sourceRows"] = [f"{entry.get('sourceSheet')}:{entry.get('sourceRow')}"]
            existing["categories"] = [entry.get("category") or "その他"]
            existing["englishVariants"] = [entry.get("en", "")]
            existing["japaneseVariants"] = [entry.get("ja", "")]
            merged[key] = existing
            continue

        existing["legacyIds"] = list(dict.fromkeys(existing["legacyIds"] + entry.get("legacyIds", [])))
        source_row = f"{entry.get('sourceSheet')}:{entry.get('sourceRow')}"
        existing["sourceRows"] = list(dict.fromkeys(existing["sourceRows"] + [source_row]))
        existing["categories"] = list(dict.fromkeys(existing["categories"] + [entry.get("category") or "その他"]))
        existing["englishVariants"] = list(dict.fromkeys(existing["englishVariants"] + [entry.get("en", "")]))
        existing["japaneseVariants"] = list(dict.fromkeys(existing["japaneseVariants"] + [entry.get("ja", "")]))
        duplicate_groups.append({
            "keptId": existing["id"],
            "canonicalEs": existing["canonicalEs"],
            "removedSource": source_row,
            "keptSource": existing["sourceRows"][0],
        })
    return list(merged.values()), duplicate_groups


def conjugation_record(sheet, row: int) -> dict:
    if "命令形" in sheet.title:
        return {
            "tu": normalize_text(sheet.cell(row=row, column=3).value) or "-",
            "usted": normalize_text(sheet.cell(row=row, column=4).value) or "-",
            "nosotros": normalize_text(sheet.cell(row=row, column=5).value) or "-",
            "ustedes": normalize_text(sheet.cell(row=row, column=6).value) or "-",
            "negativo_tu": normalize_text(sheet.cell(row=row, column=7).value) or "-",
        }
    if "分詞" in sheet.title:
        return {
            "yo": normalize_text(sheet.cell(row=row, column=3).value) or "-",
            "tu": normalize_text(sheet.cell(row=row, column=4).value) or "-",
            "el/ella": "-",
            "nosotros": "-",
            "ellos": "-",
        }
    return {
        "yo": normalize_text(sheet.cell(row=row, column=3).value) or "-",
        "tu": normalize_text(sheet.cell(row=row, column=4).value) or "-",
        "el/ella": normalize_text(sheet.cell(row=row, column=5).value) or "-",
        "nosotros": normalize_text(sheet.cell(row=row, column=6).value) or "-",
        "ellos": normalize_text(sheet.cell(row=row, column=7).value) or "-",
    }


def build_source_conjugations(wb) -> tuple[dict[str, dict], dict[str, list[str]]]:
    forms: dict[str, dict] = {}
    sources: dict[str, list[str]] = {}
    for sheet in wb.worksheets:
        tense = next((value for marker, value in CONJUGATION_SHEETS.items() if marker in sheet.title.strip()), None)
        if not tense:
            continue
        for row in range(2, sheet.max_row + 1):
            base = normalize_lexeme(sheet.cell(row=row, column=2).value)
            if not base or base in {"動詞", "スペイン語"}:
                continue
            key = search_key(base)
            forms.setdefault(key, {})[tense] = conjugation_record(sheet, row)
            sources.setdefault(key, []).append(sheet.title)
    return forms, sources


def regular_conjugations(base: str) -> dict[str, dict]:
    """Fallback for missing source rows; follows RAE regular models."""
    stem = base[:-2] if base.endswith(("ar", "er", "ir")) else base
    ending = base[-2:] if len(base) >= 2 else ""
    if ending == "ar":
        present = [stem + x for x in ("o", "as", "a", "amos", "an")]
        preterite = [stem + x for x in ("é", "aste", "ó", "amos", "aron")]
        imperfect = [stem + x for x in ("aba", "abas", "aba", "ábamos", "aban")]
        subj = [stem + x for x in ("e", "es", "e", "emos", "en")]
        imperative = [stem + "a", stem + "e", stem + "emos", stem + "en", "no " + stem + "es"]
        gerund = stem + "ando"
        participle = stem + "ado"
    elif ending == "er":
        present = [stem + x for x in ("o", "es", "e", "emos", "en")]
        preterite = [stem + x for x in ("í", "iste", "ió", "imos", "ieron")]
        imperfect = [stem + x for x in ("ía", "ías", "ía", "íamos", "ían")]
        subj = [stem + x for x in ("a", "as", "a", "amos", "an")]
        imperative = [stem + "e", stem + "a", stem + "amos", stem + "an", "no " + stem + "as"]
        gerund = stem + "iendo"
        participle = stem + "ido"
    else:
        present = [stem + x for x in ("o", "es", "e", "imos", "en")]
        preterite = [stem + x for x in ("í", "iste", "ió", "imos", "ieron")]
        imperfect = [stem + x for x in ("ía", "ías", "ía", "íamos", "ían")]
        subj = [stem + x for x in ("a", "as", "a", "amos", "an")]
        imperative = [stem + "e", stem + "a", stem + "amos", stem + "an", "no " + stem + "as"]
        gerund = stem + "iendo"
        participle = stem + "ido"
    future = [base + x for x in ("é", "ás", "á", "emos", "án")]
    conditional = [base + x for x in ("ía", "ías", "ía", "íamos", "ían")]
    result = {
        "現在形 (Presente)": dict(zip(PERSONS, present)),
        "点過去 (Indefinido)": dict(zip(PERSONS, preterite)),
        "線過去 (Imperfecto)": dict(zip(PERSONS, imperfect)),
        "未来形 (Futuro)": dict(zip(PERSONS, future)),
        "条件法 (Condicional)": dict(zip(PERSONS, conditional)),
        "接続法現在 (Subjuntivo)": dict(zip(PERSONS, subj)),
        "命令形 (Imperativo)": {
            "tu": imperative[0],
            "usted": imperative[1],
            "nosotros": imperative[2],
            "ustedes": imperative[3],
            "negativo_tu": imperative[4],
        },
        "現在・過去分詞": {
            "yo": participle,
            "tu": gerund,
            "el/ella": "-",
            "nosotros": "-",
            "ellos": "-",
        },
    }
    return apply_known_model_overrides(base, result)


def apply_known_model_overrides(base: str, forms: dict[str, dict]) -> dict[str, dict]:
    """Correct common irregular/orthographic models absent from the workbook.

    This is deliberately small and explicit.  Entries not covered here remain
    flagged for manual review in the audit manifest instead of being presented
    as fully verified RAE forms.
    """
    present_key = "現在形 (Presente)"
    preterite_key = "点過去 (Indefinido)"
    subj_key = "接続法現在 (Subjuntivo)"
    imperative_key = "命令形 (Imperativo)"
    participle_key = "現在・過去分詞"

    def set_persons(key: str, values: list[str]) -> None:
        forms[key] = dict(zip(PERSONS, values))

    if base == "caer":
        set_persons(present_key, ["caigo", "caes", "cae", "caemos", "caen"])
        set_persons(preterite_key, ["caí", "caíste", "cayó", "caímos", "cayeron"])
        set_persons(subj_key, ["caiga", "caigas", "caiga", "caigamos", "caigan"])
        forms[imperative_key] = {"tu": "cae", "usted": "caiga", "nosotros": "caigamos", "ustedes": "caigan", "negativo_tu": "no caigas"}
        forms[participle_key] = {"yo": "caído", "tu": "cayendo", "el/ella": "-", "nosotros": "-", "ellos": "-"}
    elif base in {"sentar", "tender"}:
        stem = base[:-2]
        changed = stem.replace("e", "ie", 1)
        set_persons(present_key, [changed + "o", changed + ("as" if base.endswith("ar") else "es"), changed + ("a" if base.endswith("ar") else "e"), stem + ("amos" if base.endswith("ar") else "emos"), changed + ("an" if base.endswith("ar") else "en")])
        set_persons(subj_key, [changed + ("e" if base.endswith("ar") else "a"), changed + ("es" if base.endswith("ar") else "as"), changed + ("e" if base.endswith("ar") else "a"), stem + ("emos" if base.endswith("ar") else "amos"), changed + ("en" if base.endswith("ar") else "an")])
        if base == "sentar":
            forms[imperative_key] = {"tu": "sienta", "usted": "siente", "nosotros": "sentemos", "ustedes": "sienten", "negativo_tu": "no sientes"}
        else:
            forms[imperative_key] = {"tu": "tiende", "usted": "tienda", "nosotros": "tendamos", "ustedes": "tiendan", "negativo_tu": "no tiendas"}
    elif base in {"despertar", "aprobar"}:
        stem = base[:-2]
        changed = stem.replace("o", "ue", 1)
        set_persons(present_key, [changed + "o", changed + "as", changed + "a", stem + "amos", changed + "an"])
        set_persons(subj_key, [changed + "e", changed + "es", changed + "e", stem + "emos", changed + "en"])
        if base == "despertar":
            forms[imperative_key] = {"tu": "despierta", "usted": "despierte", "nosotros": "despertemos", "ustedes": "despierten", "negativo_tu": "no despiertes"}
        else:
            forms[imperative_key] = {"tu": "aprueba", "usted": "apruebe", "nosotros": "aprobemos", "ustedes": "aprueben", "negativo_tu": "no apruebes"}
    elif base == "cruzar":
        set_persons(preterite_key, ["crucé", "cruzaste", "cruzó", "cruzamos", "cruzaron"])
        set_persons(subj_key, ["cruce", "cruces", "cruce", "crucemos", "crucen"])
        forms[imperative_key] = {"tu": "cruza", "usted": "cruce", "nosotros": "crucemos", "ustedes": "crucen", "negativo_tu": "no cruces"}
    elif base in {"acostar"}:
        set_persons(present_key, ["acuesto", "acuestas", "acuesta", "acostamos", "acuestan"])
        set_persons(subj_key, ["acueste", "acuestes", "acueste", "acostemos", "acuesten"])
        forms[imperative_key] = {"tu": "acuesta", "usted": "acueste", "nosotros": "acostemos", "ustedes": "acuesten", "negativo_tu": "no acuestes"}
    elif base in {"vestir", "medir"}:
        if base == "vestir":
            set_persons(present_key, ["visto", "vistes", "viste", "vestimos", "visten"])
            set_persons(preterite_key, ["vestí", "vestiste", "vistió", "vestimos", "vistieron"])
            set_persons(subj_key, ["vista", "vistas", "vista", "vistamos", "vistan"])
            forms[imperative_key] = {"tu": "viste", "usted": "vista", "nosotros": "vistamos", "ustedes": "vistan", "negativo_tu": "no vistas"}
            forms[participle_key] = {"yo": "vestido", "tu": "vistiendo", "el/ella": "-", "nosotros": "-", "ellos": "-"}
        else:
            set_persons(present_key, ["mido", "mides", "mide", "medimos", "miden"])
            set_persons(preterite_key, ["medí", "mediste", "midió", "medimos", "midieron"])
            set_persons(subj_key, ["mida", "midas", "mida", "midamos", "midan"])
            forms[imperative_key] = {"tu": "mide", "usted": "mida", "nosotros": "midamos", "ustedes": "midan", "negativo_tu": "no midas"}
            forms[participle_key] = {"yo": "medido", "tu": "midiendo", "el/ella": "-", "nosotros": "-", "ellos": "-"}
    elif base == "arrepentir":
        set_persons(present_key, ["arrepiento", "arrepientes", "arrepiente", "arrepentimos", "arrepienten"])
        set_persons(preterite_key, ["arrepentí", "arrepentiste", "arrepintió", "arrepentimos", "arrepintieron"])
        set_persons(subj_key, ["arrepienta", "arrepientas", "arrepienta", "arrepintamos", "arrepientan"])
        forms[imperative_key] = {"tu": "arrepiente", "usted": "arrepienta", "nosotros": "arrepintamos", "ustedes": "arrepientan", "negativo_tu": "no arrepientas"}
        forms[participle_key] = {"yo": "arrepentido", "tu": "arrepintiendo", "el/ella": "-", "nosotros": "-", "ellos": "-"}
    elif base == "fregar":
        set_persons(present_key, ["friego", "friegas", "friega", "fregamos", "friegan"])
        set_persons(preterite_key, ["fregué", "fregaste", "fregó", "fregamos", "fregaron"])
        set_persons(subj_key, ["friegue", "friegues", "friegue", "freguemos", "frieguen"])
        forms[imperative_key] = {"tu": "friega", "usted": "friegue", "nosotros": "freguemos", "ustedes": "frieguen", "negativo_tu": "no friegues"}
    elif base == "reunir":
        set_persons(present_key, ["reúno", "reúnes", "reúne", "reunimos", "reúnen"])
        set_persons(subj_key, ["reúna", "reúnas", "reúna", "reunamos", "reúnan"])
        forms[imperative_key] = {"tu": "reúne", "usted": "reúna", "nosotros": "reunamos", "ustedes": "reúnan", "negativo_tu": "no reúnas"}
    return forms


REQUESTED_TENSES = (
    "現在形 (Presente)",
    "点過去 (Indefinido)",
    "線過去 (Imperfecto)",
    "未来形 (Futuro)",
    "条件法 (Condicional)",
    "接続法現在 (Subjuntivo)",
    "命令形 (Imperativo)",
    "現在・過去分詞",
)


def complete_base_forms(base_forms: dict[str, dict] | None, base: str) -> dict[str, dict]:
    """Fill only missing tenses while preserving source forms."""
    fallback = regular_conjugations(base)
    result = {tense: dict(values) for tense, values in (base_forms or {}).items()}
    future = result.get("未来形 (Futuro)", fallback["未来形 (Futuro)"])
    if "条件法 (Condicional)" not in result:
        result["条件法 (Condicional)"] = build_conditional_from_future(future)
    for tense in REQUESTED_TENSES:
        result.setdefault(tense, dict(fallback[tense]))
    result.pop("現在完了 (Pretérito Perfecto)", None)
    return result


def append_phrase(form: str, tail: str, reflexive: bool, person: str) -> str:
    if form in ("", "-"):
        return "-"
    if reflexive and person in REFLEXIVE_PRONOUNS:
        form = f"{REFLEXIVE_PRONOUNS[person]} {form}"
    return normalize_text(f"{form} {tail}")


IMPERATIVE_CLITICS = {
    ("siente", "te"): "siéntete", ("sienta", "se"): "siéntase",
    ("sentemos", "nos"): "sentémonos", ("sientan", "se"): "siéntanse",
    ("acuesta", "te"): "acuéstate", ("acueste", "se"): "acuéstese",
    ("acostemos", "nos"): "acostémonos", ("acuesten", "se"): "acuéstense",
    ("levanta", "te"): "levántate", ("levante", "se"): "levántese",
    ("levantemos", "nos"): "levantémonos", ("levanten", "se"): "levántense",
    ("ducha", "te"): "dúchate", ("duche", "se"): "dúchese",
    ("duchemos", "nos"): "duchémonos", ("duchen", "se"): "dúchense",
    ("para", "te"): "párate", ("pare", "se"): "párese",
    ("paremos", "nos"): "parémonos", ("paren", "se"): "párense",
    ("viste", "te"): "vístete", ("vista", "se"): "vístase",
    ("vistamos", "nos"): "vistámonos", ("vistan", "se"): "vístanse",
    ("pon", "te"): "ponte", ("ponga", "se"): "póngase",
    ("pongamos", "nos"): "pongámonos", ("pongan", "se"): "pónganse",
}


def attach_clitic(form: str, pronoun: str) -> str:
    if form in ("", "-"):
        return "-"
    explicit = IMPERATIVE_CLITICS.get((form, pronoun))
    if explicit:
        return explicit
    if pronoun == "nos" and form.endswith(("amos", "emos", "imos")):
        # sentemos + nos -> sentémonos; sintamos + nos -> sintámonos
        prefix = form[:-3]
        combined = f"{prefix}monos"
        for index in range(len(prefix) - 1, -1, -1):
            if prefix[index] in "aeiou":
                accented = {"a": "á", "e": "é", "i": "í", "o": "ó", "u": "ú"}[prefix[index]]
                return combined[:index] + accented + combined[index + 1:]
        return combined

    combined = f"{form}{pronoun}"
    if pronoun in {"te", "se"} and form[-1:] in "aeo":
        # Regular affirmative imperatives: levanta + te -> levántate,
        # acostumbra + se -> acostúmbrese.
        prefix = form[:-1]
        for index in range(len(prefix) - 1, -1, -1):
            if prefix[index] in "aeiou":
                accented = {"a": "á", "e": "é", "i": "í", "o": "ó", "u": "ú"}[prefix[index]]
                return combined[:index] + accented + combined[index + 1:]
    if form.endswith(("ando", "iendo", "yendo")) and not re.search(r"[áéíóú]", form):
        prefix = combined[:-len(pronoun)]
        for index in range(len(prefix) - 3, -1, -1):
            if prefix[index] in "aeiou":
                accented = {"a": "á", "e": "é", "i": "í", "o": "ó", "u": "ú"}[prefix[index]]
                combined = prefix[:index] + accented + prefix[index + 1:] + pronoun
                break
    return combined


def phrase_conjugations(
    base_forms: dict[str, dict],
    base: str,
    tail: str,
    reflexive: bool,
    source_phrase_forms: dict[str, dict] | None = None,
) -> dict[str, dict]:
    base_forms = complete_base_forms(base_forms, base)
    result = {}
    for tense, forms in base_forms.items():
        if tense == "命令形 (Imperativo)":
            if reflexive:
                values = {
                    "tu": attach_clitic(forms.get("tu", "-"), "te"),
                    "usted": attach_clitic(forms.get("usted", "-"), "se"),
                    "nosotros": attach_clitic(forms.get("nosotros", "-"), "nos"),
                    "ustedes": attach_clitic(forms.get("ustedes", "-"), "se"),
                }
                negative = re.sub(r"^no\s+", "", forms.get("negativo_tu", "-"), flags=re.IGNORECASE)
                values["negativo_tu"] = f"no te {negative}"
            else:
                values = {
                    "tu": forms.get("tu", "-"),
                    "usted": forms.get("usted", "-"),
                    "nosotros": forms.get("nosotros", "-"),
                    "ustedes": forms.get("ustedes", "-"),
                    "negativo_tu": forms.get("negativo_tu", "-"),
                }
            result[tense] = {key: append_phrase(value, tail, False, key) for key, value in values.items()}
            continue
        if tense == "現在・過去分詞":
            phrase_participle = (source_phrase_forms or {}).get(tense)
            if phrase_participle:
                result[tense] = dict(phrase_participle)
            else:
                gerund = forms.get("tu", "-")
                if reflexive:
                    gerund = attach_clitic(gerund, "me")
                result[tense] = {
                    "yo": append_phrase(forms.get("yo", "-"), tail, False, "yo"),
                    "tu": append_phrase(gerund, tail, False, "tu"),
                    "el/ella": "-",
                    "nosotros": "-",
                    "ellos": "-",
                }
            continue
        result[tense] = {
            person: append_phrase(forms.get(person, "-"), tail, reflexive, person)
            for person in PERSONS
        }
    return result


def extract_base_and_tail(lexeme: str) -> tuple[str, str, bool]:
    tokens = lexeme.split()
    if len(tokens) == 1 and "(" not in tokens[0]:
        token = tokens[0]
        reflexive = token.endswith("se") and token[:-2].endswith(("ar", "er", "ir"))
        return (token[:-2] if reflexive else token), "", reflexive
    for index, token in enumerate(tokens):
        clean = re.sub(r"\([^)]*\)", "", token).strip(".,;:!?¡¿")
        reflexive = clean.endswith("se") and clean[:-2].endswith(("ar", "er", "ir"))
        candidate = clean[:-2] if reflexive else clean
        if candidate.endswith(("ar", "er", "ir")):
            return candidate, " ".join(tokens[index + 1:]), reflexive
    return lexeme, "", False


def build_conditional_from_future(forms: dict[str, str]) -> dict[str, str]:
    replacements = {"é": "ía", "ás": "ías", "á": "ía", "emos": "íamos", "án": "ían"}
    result = {}
    for person in PERSONS:
        value = forms.get(person, "-")
        converted = value
        for future_ending, conditional_ending in replacements.items():
            if value.endswith(future_ending):
                converted = value[: -len(future_ending)] + conditional_ending
                break
        result[person] = converted
    return result


def build_database() -> tuple[dict, dict]:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Excel source not found: {EXCEL_PATH}")
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    cefr_levels = load_cefr_levels()
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    source_forms, source_sheets = build_source_conjugations(wb)
    db = {"nouns": [], "adjectives": [], "adverbs": [], "prepositions": [], "verbs": [], "phrases": []}
    used_ids: set[str] = set()

    for sheet in wb.worksheets:
        key, part_of_speech = word_key_for_sheet(sheet.title)
        if not key:
            continue
        for row in range(2, sheet.max_row + 1):
            es_raw = sheet.cell(row=row, column=2).value
            en_raw = sheet.cell(row=row, column=3).value
            ja_raw = sheet.cell(row=row, column=4).value
            if not es_raw or not ja_raw:
                continue
            es = normalize_lexeme(es_raw)
            ja = normalize_text(ja_raw)
            en = normalize_text(en_raw)
            if is_header(es, ja):
                continue
            canonical_es, region = parse_region(es)
            legacy_id = f"{key}:{row}:{es}"
            entry = {
                "id": stable_id(key, canonical_es, region, en, ja, used_ids),
                "legacyIds": [legacy_id],
                "sourceSheet": sheet.title,
                "sourceRow": row,
                "genre": key,
                "partOfSpeech": part_of_speech,
                "category": normalize_text(sheet.cell(row=row, column=5).value) or "その他",
                "es": es,
                "canonicalEs": canonical_es,
                "ja": ja,
                "en": en,
                "region": region or None,
                "level": None,
                "frequency": None,
                "examples": [],
                "relatedExpressions": [],
                "usageScenes": [],
            }
            entry.update(audit_fields(entry, cefr_levels))
            if key == "verbs":
                base, tail, reflexive = extract_base_and_tail(canonical_es)
                entry.update({
                    "verbType": "phrase" if tail else "single",
                    "baseVerb": base,
                    "phraseTail": tail,
                    "isReflexive": reflexive,
                    "conjugations": {},
                    "conjugationAudit": {},
                })
            db[key].append(entry)

    duplicate_groups = []
    for genre_key, entries in list(db.items()):
        db[genre_key], groups = deduplicate_entries(entries)
        duplicate_groups.extend({"genre": genre_key, **group} for group in groups)

    verbs_by_base = {search_key(v["canonicalEs"]): v for v in db["verbs"] if v["verbType"] == "single"}
    active_source_keys = {
        key for key, sheets in source_sheets.items()
        if any("分詞" not in sheet_name for sheet_name in sheets)
    }
    orphan_active_source_keys = active_source_keys - set(verbs_by_base)
    conjugation_audit_counts = {"source": 0, "derived_from_base": 0, "generated_regular": 0}
    manual_review_verbs = []
    explicit_model_bases = {"caer", "sentar", "tender", "acostar", "despertar", "aprobar", "cruzar", "vestir", "medir", "arrepentir", "fregar", "reunir"}

    for verb in db["verbs"]:
        base_key = search_key(verb["baseVerb"])
        base_forms = source_forms.get(base_key)
        source_type = "source"
        if base_forms is None:
            source_type = "generated_regular"
        elif verb["verbType"] == "phrase":
            source_type = "derived_from_base"

        if verb["verbType"] == "phrase" or verb["isReflexive"]:
            verb["conjugations"] = phrase_conjugations(
                base_forms,
                verb["baseVerb"],
                verb["phraseTail"],
                verb["isReflexive"],
                source_forms.get(search_key(verb["canonicalEs"])),
            )
        else:
            verb["conjugations"] = complete_base_forms(base_forms, verb["baseVerb"])

        for tense in REQUESTED_TENSES:
            verb["conjugations"].setdefault(tense, complete_base_forms(None, verb["baseVerb"])[tense])
        verb["conjugations"].pop("現在完了 (Pretérito Perfecto)", None)
        manual_review = source_type == "generated_regular" and verb["baseVerb"] not in explicit_model_bases
        if manual_review:
            manual_review_verbs.append({"id": verb["id"], "es": verb["canonicalEs"], "baseVerb": verb["baseVerb"]})
        verb["conjugationAudit"] = {
            "status": source_type,
            "sourceSheets": source_sheets.get(base_key, []),
            "source": "RAE conjugation models + source workbook" if source_type != "derived_from_base" else "derived from RAE-reviewed base verb",
            "manualReview": manual_review,
            "reviewNote": "Workbook active-form row unavailable; regular/model fallback generated" if manual_review else None,
        }
        conjugation_audit_counts[source_type] += 1

    verb_index = {
        "single": [v["id"] for v in db["verbs"] if v["verbType"] == "single"],
        "phrase": [v["id"] for v in db["verbs"] if v["verbType"] == "phrase"],
    }
    all_entries = [entry for entries in db.values() for entry in entries]
    cefr_matched = sum(1 for entry in all_entries if entry.get("cefrLevel"))
    audit = {
        "schemaVersion": "2026-08-25",
        "sourceWorkbook": EXCEL_PATH.name,
        "counts": {key: len(values) for key, values in db.items()},
        "verbCounts": {
            "total": len(db["verbs"]),
            "single": len(verb_index["single"]),
            "phrase": len(verb_index["phrase"]),
            "withAllRequestedTenses": sum(1 for v in db["verbs"] if all(tense in v.get("conjugations", {}) for tense in REQUESTED_TENSES)),
            "orphanActiveConjugationRows": len(orphan_active_source_keys),
            "orphanActiveConjugationKeys": sorted(orphan_active_source_keys),
        },
        "conjugationAuditCounts": conjugation_audit_counts,
        "duplicateGroups": duplicate_groups,
        "manualReviewVerbs": manual_review_verbs,
        "cefrAudit": {
            "totalEntries": len(all_entries),
            "matchedToPcic": cefr_matched,
            "unmatched": len(all_entries) - cefr_matched,
            "unmatchedPolicy": "left explicit; no CEFR level guessed",
        },
        "translationAudit": {
            "totalEntries": len(all_entries),
            "workbookRetained": sum(1 for entry in all_entries if entry.get("translationStatus") == "workbook_retained"),
            "requiresBilingualReview": len(all_entries),
            "reason": "RAE/PCIC do not provide Japanese translation validation",
        },
        "sources": {
            "conjugation": "https://www.rae.es/diccionario-estudiante/docs/conjugaciones-verbales.pdf",
            "grammar": "https://www.rae.es/gtg/conjugaci%C3%B3n",
            "cefr": "https://cvc.cervantes.es/ensenanza/biblioteca_ele/plan_curricular/indice.htm",
            "cefrA1A2": "https://cvc.cervantes.es/ensenanza/biblioteca_ele/plan_curricular/niveles/09_nociones_especificas_inventario_a1-a2.htm",
            "cefrB1B2": "https://cvc.cervantes.es/ensenanza/biblioteca_ele/plan_curricular/niveles/09_nociones_especificas_inventario_b1-b2.htm",
            "cefrC1C2": "https://cvc.cervantes.es/ensenanza/biblioteca_ele/plan_curricular/niveles/09_nociones_especificas_inventario_c1-c2.htm",
        },
        "notes": [
            "The CEFR label is assigned only when the lemma matches the cached PCIC inventory; unmatched entries remain explicit rather than being guessed.",
            "Japanese and English glosses are retained from the workbook and marked for bilingual review because RAE is a Spanish-language dictionary, not a Japanese translation authority.",
            "The runtime dataset is self-contained and does not fetch fonts, images, dictionaries, or conjugations from the network.",
        ],
    }
    return db, {"verbIndex": verb_index, "audit": audit}


def write_output() -> None:
    db, metadata = build_database()
    js_content = "const db = " + json.dumps(db, ensure_ascii=False, indent=2) + ";\n\n"
    js_content += "const verbIndex = " + json.dumps(metadata["verbIndex"], ensure_ascii=False, indent=2) + ";\n\n"
    js_content += "const auditManifest = " + json.dumps(metadata["audit"], ensure_ascii=False, indent=2) + ";\n\n"
    js_content += """const genresInfo = [
  { id: "nouns", label: "名詞", enLabel: "Nouns", icon: "🏷️", color: "#FF6B6B" },
  { id: "verbs", label: "動詞", enLabel: "Verbs", icon: "🏃", color: "#3B82F6" },
  { id: "adjectives", label: "形容詞", enLabel: "Adjectives", icon: "✨", color: "#4ECDC4" },
  { id: "adverbs", label: "副詞", enLabel: "Adverbs", icon: "⏱️", color: "#FFD166" },
  { id: "prepositions", label: "その他", enLabel: "Others", icon: "🔗", color: "#8B5CF6" },
  { id: "phrases", label: "会話フレーズ", enLabel: "Phrases", icon: "💬", color: "#F59E0B" }
];"""
    JS_PATH.write_text(js_content, encoding="utf-8")
    AUDIT_MANIFEST_PATH.write_text(json.dumps(metadata["audit"], ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Excel source: {EXCEL_PATH}")
    print(f"Generated: {JS_PATH}")
    print(json.dumps(metadata["audit"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    write_output()
